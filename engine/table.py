from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from misc.util import get_temp_file_path


title_font = Font(name="Inter", bold=True)
regular_font = Font(name="Inter")


def process_input(path: Path) -> tuple[str, ...]:
    wb: Workbook = load_workbook(path)
    ws: Worksheet = wb.worksheets[0]

    # noinspection PyTypeChecker
    return tuple(map(lambda cell: cell.value, next(ws.columns)[1:]))


def process_output(contents: dict[str, list[str, ...]]) -> Path:
    wb = Workbook()
    ws: Worksheet = wb.worksheets[0]

    data = tuple(zip(*map(lambda pair: [pair[0]] + pair[1], contents.items())))
    for row in data:
        ws.append(row)

    for row in ws.rows:
        for cell in row:
            cell.font = regular_font
            cell.alignment = Alignment(wrap_text=True)

    for cell in next(ws.rows):
        cell.font = title_font

    wb.save(output_path := get_temp_file_path())

    return output_path
